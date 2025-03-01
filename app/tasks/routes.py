import io
from typing import List, Annotated

from fastapi import APIRouter, Depends, status, File, UploadFile, HTTPException, Query
from fastapi.responses import Response
from openpyxl.reader.excel import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.dependencies import AccessTokenBearer, RoleChecker
from app.db.main import get_session
from app.db.models import Task
from app.errors import TaskNotFound, InsufficientPermission
from app.tasks.schemas import TaskRead, TaskCreate, TaskUpdate
from app.tasks.service import TaskService
from app.tasks.utils import get_file_from_database

task_router = APIRouter()
task_service = TaskService()
access_token_bearer = AccessTokenBearer()
role_checker = Depends(RoleChecker(['admin', 'user']))


VALID_CODE = '202502'
DOWNLOAD_APK_URL = f"https://firebasestorage.googleapis.com/v0/b/dagenergi-b0086.appspot.com/o/apk%2Fapp-release.apk.zip?alt=media&token=248b1700-a781-45d5-99db-44ffe94d7048"


@task_router.get("/download_apk")
async def download_apk(
		code: str = Query(..., min_length=6, max_length=6)
):
	if code != VALID_CODE:
		raise HTTPException(status_code=403, detail="Invalid Code")
	return  {"download_url": DOWNLOAD_APK_URL}

@task_router.get("/", response_model=List[Task], dependencies=[role_checker])
async def get_all_tasks(
		session: AsyncSession = Depends(get_session),
		_: dict = Depends(access_token_bearer)
):
	tasks = await task_service.get_all_tasks(session)
	return tasks

@task_router.get("/completed", response_model=List[Task])
async def get_completed_task(session: Annotated[AsyncSession, Depends(get_session)]):
	completed_tasks = await task_service.get_tasks_completed(session)
	return completed_tasks

@task_router.get("/{task_id}", response_model=TaskRead, dependencies=[role_checker])
async def get_task(
		task_id: int,
		session: AsyncSession = Depends(get_session),
		_: dict = Depends(access_token_bearer)
) -> dict:
	task = await task_service.get_task(task_id, session)

	if task:
		return task
	else:
		raise TaskNotFound


@task_router.post(
	"/",
	status_code=status.HTTP_201_CREATED,
	response_model=TaskRead,
	dependencies=[role_checker]
)
async def create_a_task(
		task_data: TaskCreate,
		session: AsyncSession = Depends(get_session),
		token_details: dict = Depends(access_token_bearer)
):
	username = token_details.get("user")["username"]
	new_task = await task_service.create_a_task(task_data, username, session)
	return new_task


@task_router.post(
	"/upload",
	status_code=status.HTTP_201_CREATED,
	response_model=List[TaskRead]
)
async def upload_file(
		uploadFile: UploadFile = File(...),
		session: AsyncSession = Depends(get_session),
):
	uploadFile.file.seek(0)
	content = uploadFile.file.read()
	workbook = load_workbook(io.BytesIO(content))
	sheet = workbook.active
	tasks = []
	for row in range(3, sheet.max_row + 1):
		try:
			new_task = TaskCreate(
				work_type=str(sheet.cell(row=row, column=2).value) if sheet.cell(row=row, column=2).value else None,
				dispatcher_name=str(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value else None,
				address=str(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value else None,
				planner_date=str(sheet.cell(row=row, column=5).value) if sheet.cell(row=row, column=5).value else None,
				voltage=sheet.cell(row=row, column=7).value if sheet.cell(row=row, column=7).value else None,
				job=str(sheet.cell(row=row, column=8).value) if sheet.cell(row=row, column=8).value else None,
				latitude=None,
				longitude=None,
				photo_url_1=None,
				photo_url_2=None,
				photo_url_3=None,
				photo_url_4=None,
				photo_url_5=None,
				comments=None
			)
		except KeyError as e:
			raise HTTPException(
				status_code=400, detail=f"Missing column in the Excel file: {e}"
			)
		task = await task_service.create_task_from_file(new_task, session)
		tasks.append(task)
	return tasks


@task_router.patch(
	"/{task_id}",
	response_model=TaskRead,
	dependencies=[role_checker]
)
async def update_task(
		task_id: int, task_update_data: TaskUpdate, session: AsyncSession = Depends(get_session),
		token_details: dict = Depends(access_token_bearer)
):
	username = token_details.get('user')["username"]
	updated_task = await task_service.update_task(task_id, task_update_data, session, username)
	if updated_task is None:
		raise TaskNotFound
	else:
		return updated_task


@task_router.delete(
	"/clear", status_code=status.HTTP_204_NO_CONTENT
)
async def delete_all_tasks(
		session: AsyncSession = Depends(get_session)
):
	all_tasks = await task_service.tasks_delete(session)
	return all_tasks


@task_router.delete(
	"/{task_id}", status_code=status.HTTP_204_NO_CONTENT, dependencies=[role_checker]
)
async def delete_task(
		task_id: int, session: AsyncSession = Depends(get_session),
		token_data: dict = Depends(access_token_bearer)
):
	role = token_data.get('user')['role']
	if role == 'admin':
		task_to_delete = await task_service.task_delete(task_id, session)

		if task_to_delete is None:
			raise TaskNotFound
		else:
			return {}
	else:
		raise InsufficientPermission()


@task_router.post("/download", status_code=status.HTTP_201_CREATED)
async def download(
	session: AsyncSession = Depends(get_session),
):
	tasks = await task_service.get_tasks_completed(session)
	file = get_file_from_database(tasks)
	file_content = file.getvalue()
	headers = {
		'Content-Disposition': 'attachment; filename="Reports.xlsx"',
		"Access-Control-Allow-Origin": "*",
		"Access-Control-Allow-Headers": "*",
		"Access-Control_Allow-Methods": "POST, GET, OPTIONS",
	}
	return Response(content=file_content,
					media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8",
					headers=headers)

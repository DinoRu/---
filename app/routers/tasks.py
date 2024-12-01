import io
import uuid
from typing import List

from fastapi import APIRouter, status, Depends, HTTPException, File, UploadFile
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi.responses import Response

from app.authentication import get_current_user
from app.controllers.tasks import task_controller
from app.database import get_session
from app.models.users import User
from app.schemas.tasks import TaskComplete, CreateTask, TaskUpdate, AddNewTask
from app.schemas.users import UserOut
from app.utils.status import TaskStatus

router = APIRouter(
	prefix="/tasks",
	tags=["Tasks"],
	responses={404: {"description": "Not found."}}
)


@router.post("/add_task",
			 status_code=status.HTTP_201_CREATED,
			 response_model=TaskComplete,
			 summary="Add new task.")
async def add_task(
		data_schema: AddNewTask,
		user: User = Depends(get_current_user),
		session: AsyncSession = Depends(get_session),
):
	task = await task_controller.add_new_task(
		session=session,
		data=data_schema,
		supervisor=user.full_name
	)
	return task

@router.post("/upload",
			 status_code=status.HTTP_201_CREATED,
			 response_model=List[TaskComplete],
			 summary="Upload file (.xlsx)")
async def import_tasks_from_excel(
		file: UploadFile = File(...),
		session: AsyncSession = Depends(get_session)
):
	# if not file.filename.endswith(".xlsx"):
	# 	raise HTTPException(
	# 		status_code=status.HTTP_400_BAD_REQUEST,
	# 		detail="Only .xlsx files are supported."
	# 	)
	# Load the Excel file
	file.file.seek(0)
	content = file.file.read()
	workbook = load_workbook(io.BytesIO(content))
	sheet = workbook.active
	tasks = []
	for row in range(2, sheet.max_row + 1):
		try:
			task_data = CreateTask(
				task_id=uuid.uuid4(),
				code=str(sheet.cell(row=row, column=1).value) if sheet.cell(row=row, column=1).value else "",
				dispatcher_name=str(sheet.cell(row=row, column=2).value) if sheet.cell(row=row, column=2).value else "",
				location=str(sheet.cell(row=row, column=3).value) if sheet.cell(row=row, column=3).value else "",
				planner_date=str(sheet.cell(row=row, column=4).value) if sheet.cell(row=row, column=4).value else "",
				voltage_class=float(sheet.cell(row=row, column=5).value) if sheet.cell(row=row, column=5).value else 0.0,
				work_type=str(sheet.cell(row=row, column=6).value) if sheet.cell(row=row, column=6).value else "",
				completion_date=None,
				latitude=None,
				longitude=None,
				photo_url_1=None,
				photo_url_2=None,
				supervisor=None,
				comments=None,
				status=TaskStatus.EXECUTING
			)
		except KeyError as e:
			raise HTTPException(
				status_code=400, detail=f"Missing column in the Excel file: {e}"
			)
		task = await task_controller.add_task(session=session, data=task_data)
		tasks.append(task)
	total = len(tasks)
	print(total)
	return tasks


@router.put("/task/{task_id}",
			status_code=status.HTTP_200_OK,
			summary="Update task.",
			response_model=TaskComplete)
async def completed_task(
		task_id: uuid.UUID,
		data: TaskUpdate,
		session: AsyncSession = Depends(get_session),
		user: UserOut = Depends(get_current_user)
):
	task = await task_controller.modify_task(
		session=session,
		task_id=task_id,
		update_data=data,
		username=user.full_name
	)
	return task

@router.get("/task_by/{task_status/",
			status_code=status.HTTP_200_OK,
			summary="Get task by status.",
			response_model=List[TaskComplete])
async def get_tasks_by_status(
		task_status: TaskStatus,
		session: AsyncSession = Depends(get_session)
):
	tasks = await task_controller.get_by_status(session, task_status)
	return tasks

@router.get("/tasks/completed/",
			status_code=status.HTTP_200_OK,
			response_model=List[TaskComplete],
			summary="Get completed tasks.")
async def completed_tasks(session: AsyncSession = Depends(get_session)):
	tasks = await task_controller.get_completed_tasks(session)
	return tasks

@router.get("/tasks/pending/",
			status_code=status.HTTP_200_OK,
			response_model=List[TaskComplete],
			summary="Tasks pending tasks.")
async def get_pending_tasks(session: AsyncSession = Depends(get_session)):
	pending_tasks = await task_controller.get_pending_tasks(session)
	return pending_tasks


@router.get("/tasks/supervisor", status_code=status.HTTP_200_OK, summary="Get tasks completed by supervisor.")
async def tasks_completed_by_supervisor(
		supervisor: User = Depends(get_current_user),
		session: AsyncSession = Depends(get_session)):
	tasks = await task_controller.get_completed_tasks_by_supervisor(session, supervisor.full_name)
	return tasks


@router.get("/supervisor/tasks", status_code=status.HTTP_200_OK, summary="Get tasks assigned by supervisor.")
async def assigned_tasks_by_user(
		session: AsyncSession = Depends(get_session),
		supervisor: User = Depends(get_current_user)
):
	tasks = await task_controller.get_tasks_by_user(
		session=session,
		location=supervisor.location
	)
	return tasks

@router.get("/completed/tasks", status_code=status.HTTP_200_OK, summary="Get tasks completed by supervisor.")
async def completed_tasks_by_user(
		session: AsyncSession = Depends(get_session),
		supervisor: User = Depends(get_current_user)
):
	tasks = await task_controller.get_tasks_completed_by_user(
		session=session,
		location=supervisor.location
	)
	return tasks

@router.get("/",
			status_code=status.HTTP_200_OK,
			summary="Get all tasks",
			response_model=List[TaskComplete]
			)
async def get_all_tasks(
		session: AsyncSession = Depends(get_session)
):
	return await task_controller.all(session=session)

@router.get("/task/{task_id}",
			status_code=status.HTTP_200_OK,
			summary="Get Task by task_ID.",
			response_model=TaskComplete)
async def get_one_or_404(
		task_id: uuid.UUID,
		session: AsyncSession = Depends(get_session)
):
	return await task_controller.get_task_or_404(session, task_id)

@router.delete("/",
			   status_code=status.HTTP_200_OK,
			   summary="Delete all Tasks.")
async def remove_all_tasks(
		session: AsyncSession = Depends(get_session)
):
	return await task_controller.delete_all(session=session)


@router.delete("/task/{task_id}",
			   status_code=status.HTTP_200_OK,
			   summary="Remove task by Task_ID.")
async def remove_task(
		task_id: uuid.UUID,
		session: AsyncSession = Depends(get_session)
):
	return await task_controller.delete_or_404(session=session, task_id=task_id)

@router.post('/download',
			 status_code=status.HTTP_201_CREATED,
			 summary="Download completed tasks",
			 response_description="Excel file with completed tasks"
			 )
async def download_completed_tasks(session: AsyncSession = Depends(get_session)):
	file_stream = await task_controller.get_completed_tasks_files(session=session)
	file_content = file_stream.getvalue()
	headers = {
		'Content-Disposition': 'attachment; filename="Reports.xlsx"',
		"Access-Control-Allow-Origin": "*",
		"Access-Control-Allow-Headers": "*",
		"Access-Control_Allow-Methods": "POST, GET, OPTIONS",
	}
	return Response(content=file_content,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8",
                    headers=headers)

